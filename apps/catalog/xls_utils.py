# -*- coding: utf-8 -*-
from __future__ import unicode_literals

from django.utils import timezone
from django.utils.html import strip_tags

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment

from .models import Product, ProductOption, ProductExtraOption


def dump_catalog_products(lang='en', lang_label='English'):
    book = Workbook()

    BOLD = Font(bold=True)
    CENTER = Alignment(horizontal='center', vertical='center')
    WRAP_TEXT = Alignment(wrap_text=True)

    # создаем две вкладки
    s1 = book.active
    s1.title = 'Заголовки'
    s2 = book.create_sheet('Тексты')

    # заголовки столбцов
    for s in [s1, s2]:
        s.append(('ID', 'Russian', lang_label))
        for id in ['A1', 'B1', 'C1']:
            s[id].font = BOLD
            s[id].alignment = CENTER

    # пишем тексты товаров, вариантов и доп.товаров
    products = Product.objects.all().prefetch_related('options', 'extra_options').order_by('id')
    for p in products:
        _id = '{}t'.format(p.id)
        _ru = p.title_ru or ''
        _xx = getattr(p, 'title_{}'.format(lang), '') or ''
        s1.append((_id, _ru, _xx))

        if p.subtitle_ru:
            _id = '{}s'.format(p.id)
            _ru = p.subtitle_ru or ''
            _xx = getattr(p, 'subtitle_{}'.format(lang), '') or ''
            s1.append((_id, _ru, _xx))

        for o in p.options.all():
            if o.title_ru:
                _id = '{}_{}o'.format(p.id, o.id)
                _ru = o.title_ru or ''
                _xx = getattr(o, 'title_{}'.format(lang), '') or ''
                s1.append((_id, _ru, _xx))

        for e in p.extra_options.all():
            if e.title_ru:
                _id = '{}_{}e'.format(p.id, e.id)
                _ru = e.title_ru or ''
                _xx = getattr(e, 'title_{}'.format(lang), '') or ''
                s1.append((_id, _ru, _xx))

        if p.text_ru:
            _id = '{}t'.format(p.id)
            _ru = strip_tags(p.text_ru or '').replace('&nbsp;', ' ')
            _text = getattr(p, 'text_{}'.format(lang), '') or ''
            _xx = strip_tags(_text).replace('&nbsp;', ' ')
            s2.append((_id, _ru, _xx))

    # ширина столбцов
    s1.column_dimensions['B'].width = 60
    s1.column_dimensions['C'].width = 60
    s2.column_dimensions['C'].width = 60
    s2.column_dimensions['B'].width = 60

    # высота строк
    for i in xrange(2, s2.max_row+1):
        # s2.row_dimensions[i].height = 100
        s2['B{}'.format(i)].alignment = WRAP_TEXT
        s2['C{}'.format(i)].alignment = WRAP_TEXT

    # пишем в файл
    now_str = timezone.now().strftime('%Y%m%d_%H%M%S')
    filename = 'static/temp/{}_catalog_{}.xlsx'.format(now_str, lang)
    book.save(filename)
    book.close()
    print filename


def load_catalog(sheets=None, filename='apps/catalog/tempo/catalog_en.xlsx', lang='en'):
    if sheets is None:
        sheets = [1, 2]
    book = load_workbook(filename)

    if 1 in sheets:
        # заголовки
        s1 = book.worksheets[0]
        for row in s1:
            _id = row[0].value
            _title = row[2].value

            try:
                if _id.endswith('t'):
                    product_id = _id[:-1]
                    product = Product.objects.get(id=product_id)
                    setattr(product, 'title_{}'.format(lang), _title)
                    product.save()

                elif _id.endswith('s'):
                    product_id = _id[:-1]
                    product = Product.objects.get(id=product_id)
                    setattr(product, 'subtitle_{}'.format(lang), _title)
                    product.save()

                elif _id.endswith('o'):
                    o_id = _id[:-1]
                    product_id, option_id = o_id.split('_')
                    option = ProductOption.objects.get(id=option_id, product_id=product_id)
                    setattr(option, 'title_{}'.format(lang), _title)
                    option.save()

                elif _id.endswith('e'):
                    e_id = _id[:-1]
                    product_id, option_id = e_id.split('_')
                    option = ProductExtraOption.objects.get(id=option_id, product_id=product_id)
                    setattr(option, 'title_{}'.format(lang), _title)
                    option.save()

                else:
                    raise ValueError('Error in ID')

            except (ValueError, Product.DoesNotExist, ProductOption.DoesNotExist, ProductExtraOption.DoesNotExist) as exc:
                print '{} "{}" | {}'.format(_id, _title, exc)

            else:
                print '{} "{}" | DONE'.format(_id, _title)

        print '1 done!'; print; print

    if 2 in sheets:
        # тексты
        s2 = book.worksheets[1]
        for row in s2:
            _id = row[0].value
            _text = row[2].value

            if _text:
                try:
                    if _id.endswith('t'):
                        product_id = _id[:-1]
                        product = Product.objects.get(id=product_id)
                        _text = '<p>{}</p>'.format('</p>\n\n<p>'.join(_text.split('\n\n'))).replace('\n', '\r\n')
                        setattr(product, 'text_{}'.format(lang), _text)
                        product.save()
                    else:
                        raise ValueError('Error in ID')

                except (ValueError, Product.DoesNotExist) as exc:
                    print '{} | {}'.format(_id, exc)

                else:
                    print '{} | DONE'.format(_id)
        print '2 done!'; print; print

    print 'done'
